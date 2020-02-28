from openpyxl import Workbook, load_workbook
# wb = Workbook()
# ws = wb.active
# ws.title = 'rowData'
# ws['A1'] = "rowNum"
# ws['B1'] = "name"
# wb.save('dataAnalysis.xlsx')

# c = ws['A1']
# print(c.value)

# 读取要整理格式的数据
raw = load_workbook('rawData.xlsx')
rawData = raw['Sheet1']

a = 2
b = 'A'+str(a)
# print(b)
# print(rawData[b])
# print(rawData[b].value)
# print(rawData['A'+str(a)].value)

# 字段拼接
for i in range(2, 10000000000):
    if rawData['F'+str(i)].value == None:
        break
    rawData['J'+str(i)] = str(rawData['F'+str(i)].value) + \
        '('+str(rawData['G'+str(i)].value) + \
        ','+str(rawData['H'+str(i)].value)+')'
    print(rawData['J'+str(i)].value)

# 单元格内容切分
# 根据统一标识切分

# 根据固定长度切分


raw.save('afterManagement.xlsx')

# print(rawData['A'+str(a)])
