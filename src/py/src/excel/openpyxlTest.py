# 引入第三方库openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# 打开已存在的表格
wb2 = load_workbook('forOpen.xlsx')
sheet = wb2.active
sheet.append(['测试插入数据', 123])
wb2.save('forOpen.xlsx')

# 实例化
wb = Workbook()
# 激活worksheet
ws = wb.active
# 重命名
ws.title = "New Title"
# 直接赋值
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['A4'] = '=SUM(A1:A3)'
# 在下方空行从左往右填入list中的内容
ws.append([1, 2, 3, 4])
# 创建表格,指定排序时（0开始），要注意不要指定以有表格的序列，会覆盖原工作表
wb.create_sheet("my_new_sheet", 1)
wb.create_sheet("my_new_sheet2", 2)
wb.create_sheet("my_new_sheet3")  # 不指定的话默认往后加

# 选择表
ws3 = wb["my_new_sheet"]
ws4 = wb.get_sheet_by_name("my_new_sheet3")
print(ws4.title)

print(wb.sheetnames)
ws4['B2'] = 'B2'

# 保存表格
wb.save("test.xlsx")
